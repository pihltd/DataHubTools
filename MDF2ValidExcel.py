import yaml
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter, quote_sheetname
import argparse

def readYaml(yamlfile):
    with open(yamlfile) as f:
        jsonstuff = yaml.load(f, Loader=yaml.FullLoader)
    return jsonstuff

def validationColumn(wb, sheetname, prop, values, valcol):
    print(wb.sheetnames)
    sheet = wb.get_sheet_by_name(sheetname)
    row = 1
    sheet.cell(row=row, column=valcol).value = prop
    for value in values:
        row = row + 1
        sheet.cell(row=row, column=valcol).value = value

def main(args):
    modelyaml = args.modelfile
    propyaml = args.propsfile

    model = readYaml(modelyaml)
    props = readYaml(propyaml)


    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Validations")

    valcol = 1

    for node, nodeinfo in model['Nodes'].items():
        row = 1
        col = 1
        proplist = nodeinfo['Props']
        ws = wb.create_sheet(node)
        enums = {}
        for prop in proplist:
            #Write the column header
            ws.cell(row=row, column=col).value = prop
            if(args.validation):
                #Check and see if the property has enumerated values
                #
                #https://medium.com/@berkeozbek1997/breaking-the-256-character-limit-adding-a-dropdown-list-to-excel-e43e24ef432#:~:text=If%20you've%20ever%20worked,options%20or%20complex%20data%20structures.
                # 
                #So the problem with validations is providing them as a list is limited to 256 characters in the list.  The only way around that appears to be to put the values in a column and then reference the column
                
                if 'Enum' in props['PropDefinitions'][prop]:
                    #There are enums let's write the list to a column (gets around the 256 character limit)
                    #Set up the info needed to write out the enum list and to create teh validation.  Yeah, this could be done in the statements, but this is easier to understand
                    enumlist = props['PropDefinitions'][prop]['Enum']
                    #Valcoolumn is tracking the column on the validation sheet.  Column is tracking the column on the node loading sheet.
                    valcolumn = get_column_letter(valcol)
                    column = get_column_letter(col)
                    enumlength = len(enumlist)
                    #Add a column of validations to the Validations sheet
                    validationColumn(wb,'Validations', prop, enumlist, valcol)
                    #Set up the validtion
                    valsheet = quote_sheetname("Validations")
                    valrange = f"{valsheet}!{valcolumn}2:{valcolumn}{enumlength}"
                    dv = DataValidation(type='list', formula1=valrange, allow_blank=True, showDropDown=False)
                    #The question is how many rows to have the drop-down pre-populated.  Usign 10 for testing purposes
                    ##maxrow = 1048576
                    maxrow = 10
                    colrange = "{}2:{}{}".format(column, column, str(maxrow))
                    dv.add(colrange)
                    ws.add_data_validation(dv)
                    valcol = valcol + 1

            col = col +1
    wb.save(args.outputfile)
    wb.close()
        
if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("-m", "--modelfile", required=True, help="MDF model file")
    parser.add_argument("-p", "--propsfile", required=True, help="MDF Properties file")
    parser.add_argument("-o", "--outputfile", required=True, help="Output spreadsheet")
    parser.add_argument("-v", "--validation", action='store_true', help="Create workbook with column validation")
        
    args = parser.parse_args()
    main(args)
    